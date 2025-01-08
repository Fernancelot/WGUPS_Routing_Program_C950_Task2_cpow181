# main.py
# Author: Christopher Powell
# Student ID: 001307071

from datetime import datetime, time, timedelta
import openpyxl
from typing import Dict, List, Optional, Set
from package import Package, PackageError
from truck import Truck, TruckError
from hash_table import HashTable

class DeliveryServiceError(Exception):
    pass

class DeliveryService:
    def __init__(self):
        self.packages = HashTable()
        self.distances: Dict[str, Dict[str, float]] = {}
        self.addresses: List[str] = []
        self.trucks = [
            Truck(1),
            Truck(2),
            Truck(3)
        ]

    def load_data(self, package_file: str, distance_file: str) -> None:
        try:
            self._load_packages(package_file)
            self._load_distances(distance_file)
        except Exception as e:
            raise DeliveryServiceError(f"Failed to load data: {str(e)}")

    def _load_packages(self, filename: str) -> None:
        try:
            wb = openpyxl.load_workbook(filename)
            sheet = wb.active
            
            for row in list(sheet.iter_rows(min_row=2, values_only=True)):
                if not row[0]:
                    continue
                    
                package = Package(
                    package_id=row[0],
                    address=row[1],
                    city=row[2],
                    state=row[3],
                    zip_code=row[4],
                    deadline=row[5],
                    weight=row[6],
                    notes=row[7] if row[7] else ""
                )
                self.packages.insert(package.package_id, package)
        except Exception as e:
            raise DeliveryServiceError(f"Error loading packages: {str(e)}")

    def _load_distances(self, filename: str) -> None:
        try:
            wb = openpyxl.load_workbook(filename)
            sheet = wb.active
            
            # Get the address row, skip first column which is empty
            rows = list(sheet.iter_rows(values_only=True))
            self.addresses = []
            
            # Process header row to get clean addresses
            for addr in rows[1][1:]:  # Skip first empty cell
                if addr:
                    # Extract just the address part before any newlines or parentheses
                    clean_addr = str(addr).split('\n')[0].strip()
                    self.addresses.append(clean_addr)
            
            # Process distance data
            for row in rows[2:]:
                if not row[0]:
                    continue
                    
                # Clean up the from_address
                from_addr = str(row[0]).split('\n')[0].strip()
                
                # Create distance mapping
                self.distances[from_addr] = {}
                
                for i, dist in enumerate(row[1:]):
                    if dist and i < len(self.addresses):
                        try:
                            # Convert to float if it's a number
                            if isinstance(dist, (int, float)) or (isinstance(dist, str) and dist.strip().replace('.', '').isdigit()):
                                self.distances[from_addr][self.addresses[i]] = float(dist)
                        except ValueError:
                            continue  # Skip if conversion fails
        except Exception as e:
            raise DeliveryServiceError(f"Error loading distances: {str(e)}")

    def get_distance(self, addr1: str, addr2: str) -> float:
        addr1 = str(addr1).strip()
        addr2 = str(addr2).strip()
        
        try:
            return self.distances[addr1][addr2]
        except KeyError:
            try:
                return self.distances[addr2][addr1]
            except KeyError:
                raise DeliveryServiceError(f"Distance not found between {addr1} and {addr2}")

    def optimize_delivery_routes(self) -> None:
        for truck in self.trucks:
            if truck.packages:
                truck.optimize_route(self.distances)

    def deliver_packages(self) -> None:
        # Set departure times
        departure_times = [
            datetime.combine(datetime.today(), time(8, 0)),    # Truck 1: 8:00 AM
            datetime.combine(datetime.today(), time(9, 5)),    # Truck 2: 9:05 AM
            datetime.combine(datetime.today(), time(10, 20))   # Truck 3: 10:20 AM
        ]
        
        # Load and deliver packages for each truck
        self._load_trucks(departure_times)
        self.optimize_delivery_routes()
        
        for truck in self.trucks:
            if truck.packages:
                self._deliver_truck_packages(truck)

    def _load_trucks(self, departure_times: List[datetime]) -> None:
        packages = self.packages.get_all()
        
        # Sort packages by constraints and priority
        truck2_required: List[Package] = []
        delayed_packages: List[Package] = []
        priority_packages: List[Package] = []
        grouped_packages: List[Package] = []
        regular_packages: List[Package] = []
        
        for package in packages:
            package_id = package.package_id
            
            if package_id in {3, 18, 36, 38}:
                truck2_required.append(package)
            elif package_id in {6, 25, 28, 32}:
                delayed_packages.append(package)
            elif package_id in {13, 14, 15, 16, 19, 20}:
                grouped_packages.append(package)
            elif package.deadline != time(17, 0):
                priority_packages.append(package)
            else:
                regular_packages.append(package)

        # Load truck 2 (special requirements)
        truck2_load = truck2_required + delayed_packages
        self.trucks[1].load_packages(truck2_load, departure_times[1])
        
        # Load truck 1 (grouped and priority packages)
        truck1_load = grouped_packages + [p for p in priority_packages 
                                        if p not in truck2_load][:16-len(grouped_packages)]
        self.trucks[0].load_packages(truck1_load, departure_times[0])
        
        # Load truck 3 (remaining packages)
        remaining = [p for p in regular_packages + priority_packages 
                    if p not in truck2_load and p not in truck1_load]
        self.trucks[2].load_packages(remaining[:16], departure_times[2])

    def _deliver_truck_packages(self, truck: Truck) -> None:
        while truck.packages:
            next_package = self._find_nearest_package(truck)
            distance = self.get_distance(truck.current_location, next_package.address)
            delivery_time = truck.deliver_package(next_package, distance)
            next_package.update_status("Delivered", delivery_time)
        
        # Return to hub
        distance = self.get_distance(truck.current_location, "HUB")
        truck.return_to_hub(distance)

    def _find_nearest_package(self, truck: Truck) -> Package:
        if not truck.packages:
            raise DeliveryServiceError("No packages left on truck")
            
        nearest_package = min(truck.packages,
            key=lambda p: self.get_distance(truck.current_location, p.address))
        return nearest_package

    def get_package_status(self, package_id: int, check_time: datetime) -> Optional[dict]:
        package = self.packages.lookup(package_id)
        if package:
            return package.get_status_at_time(check_time)
        return None

    def get_all_package_status(self, check_time: datetime) -> List[dict]:
        return [p.get_status_at_time(check_time) for p in self.packages.get_all()]

    def get_total_mileage(self) -> float:
        return sum(truck.mileage for truck in self.trucks)

def main():
    service = DeliveryService()
    
    try:
        print("\nInitializing WGUPS delivery service...")
        service.load_data('WGUPS Package File.xlsx', 'WGUPS Distance Table.xlsx')
        
        print("Optimizing and delivering packages...")
        service.deliver_packages()
        
        while True:
            print("\nWGUPS Package Tracking System")
            print("1. Check all package status at a specific time")
            print("2. Check specific package status")
            print("3. View total mileage")
            print("4. Exit")
            
            try:
                choice = input("\nSelect an option: ")
                
                if choice == '1':
                    time_str = input("Enter time (HH:MM AM/PM): ")
                    check_time = datetime.combine(
                        datetime.today(),
                        datetime.strptime(time_str, "%I:%M %p").time()
                    )
                    
                    statuses = service.get_all_package_status(check_time)
                    print("\nPackage Status Report:")
                    print("-" * 80)
                    for status in statuses:
                        print(
                            f"Package {status['id']}: {status['status']}\n"
                            f"  Address: {status['address']}, {status['city']}, "
                            f"{status['state']} {status['zip']}\n"
                            f"  Deadline: {status['deadline']}, Weight: {status['weight']} KILO"
                        )
                    print("-" * 80)
                    
                elif choice == '2':
                    package_id = int(input("Enter package ID: "))
                    time_str = input("Enter time (HH:MM AM/PM): ")
                    check_time = datetime.combine(
                        datetime.today(),
                        datetime.strptime(time_str, "%I:%M %p").time()
                    )
                    
                    status = service.get_package_status(package_id, check_time)
                    if status:
                        print("\nPackage Status:")
                        print("-" * 80)
                        print(
                            f"Package {status['id']}: {status['status']}\n"
                            f"Address: {status['address']}, {status['city']}, "
                            f"{status['state']} {status['zip']}\n"
                            f"Deadline: {status['deadline']}, "
                            f"Weight: {status['weight']} KILO"
                        )
                        if status['notes']:
                            print(f"Notes: {status['notes']}")
                        print("-" * 80)
                    else:
                        print(f"Package {package_id} not found")
                        
                elif choice == '3':
                    total_mileage = service.get_total_mileage()
                    print(f"\nTotal mileage for all trucks: {total_mileage:.1f} miles")
                    if total_mileage < 140:
                        print("✓ Total mileage is under the required 140 mile limit")
                    else:
                        print("! Warning: Total mileage exceeds 140 mile limit")
                        
                elif choice == '4':
                    print("\nThank you for using WGUPS Package Tracking System")
                    break
                    
                else:
                    print("Invalid option. Please try again.")
                    
            except ValueError as e:
                print(f"Invalid input: {str(e)}")
            except Exception as e:
                print(f"An error occurred: {str(e)}")
                
    except Exception as e:
        print(f"Fatal error: {str(e)}")
        return 1

if __name__ == "__main__":
    main()